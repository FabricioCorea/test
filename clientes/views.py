from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User, Group
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from datetime import date, datetime
from django.utils.timezone import localdate
from clientes.models import *
from django.utils import timezone
import pytz
from django.db.models import Count
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from django.db.models import Max, F, Prefetch
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Exists, OuterRef, Q
from django.db.models import Subquery,  DateTimeField
from collections import defaultdict
from django.core.paginator import Paginator


@login_required
def index(request):
    return render(request, 'inicio/inicio.html')

def paginar_queryset(request, queryset, param, por_pagina=10):
    paginator = Paginator(queryset, por_pagina)
    pagina = request.GET.get(param)
    return paginator.get_page(pagina)

@login_required
def clientes_pendientes(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get('q', '').strip()

    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
 
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    clientes_pendientes_qs = Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_pendiente)
    if search_query:
        clientes_pendientes_qs = clientes_pendientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes_pendientes = paginar_queryset(request, clientes_pendientes_qs, 'pendientes')

    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    estados_reporte = EstadoReporte.objects.filter(estado="activo") \
        .exclude(nombre__iexact="pendiente") \
        .exclude(nombre__iexact="no localizado") \
        .exclude(nombre__iexact="por localizar") \
        .exclude(nombre__iexact="formulario sin respuesta")
    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()


    return render(request, 'clientes/clientes.html', {
        "clientes_pendientes": clientes_pendientes,
        "estado_reporte": estados_reporte,
        "view_type": "pendientes",
        "search_query": search_query,
        "count_pendientes": clientes_pendientes_qs.count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(estado_actual__nombre__iexact="actualizado", movimientos__estado__nombre__iexact="actualizado", movimientos__actualizado_por=usuario).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_seguimiento(request):
    hoy = timezone.localdate() 
    usuario = request.user

    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()

    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )

    if search_query:
        clientes_seguimiento_qs = clientes_seguimiento_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes_seguimiento = paginar_queryset(request, clientes_seguimiento_qs, 'seguimiento')

    # -------- Cálculo de clientes sin actualizar ----------
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    estados_reporte = EstadoReporte.objects.filter(estado="activo") \
        .exclude(nombre__iexact="pendiente") \
        .exclude(nombre__iexact="no localizado") \
        .exclude(nombre__iexact="por localizar") \
        .exclude(nombre__iexact="formulario sin respuesta")
    
    # -------- Clientes actualizados hoy por el usuario ----------
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()

    return render(request, 'clientes/clientes.html', {
        "clientes_seguimiento": clientes_seguimiento,
        "estado_reporte": estados_reporte,
        "estado_seguimiento": EstadoReporte.objects.filter(nombre__iexact="formulario enviado").first(),
        "view_type": "seguimiento",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado",
            movimientos__actualizado_por=usuario
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_sin_contestar(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    clientes_sin_contestar_qs = Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_no_contesto)
    if search_query:
        clientes_sin_contestar_qs = clientes_sin_contestar_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )
    clientes_sin_contestar = paginar_queryset(request, clientes_sin_contestar_qs, 'nocontesto')

    # Contador sin actualizar
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    # Estados disponibles, excluyendo los no deseados
    estados_reporte = EstadoReporte.objects.filter(estado="activo") \
        .exclude(nombre__iexact="pendiente") \
        .exclude(nombre__iexact="no localizado") \
        .exclude(nombre__iexact="por localizar") \
        .exclude(nombre__iexact="formulario sin respuesta")
    
    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()
    
    return render(request, 'clientes/clientes.html', {
        "clientes_sin_contestar": clientes_sin_contestar,
        "estado_reporte": estados_reporte,
        "view_type": "nocontesto",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(estado_actual__nombre__iexact="actualizado", movimientos__estado__nombre__iexact="actualizado", movimientos__actualizado_por=usuario).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_sin_actualizar(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_filtrados_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).select_related('estado_actual')

    if search_query:
        clientes_filtrados_qs = clientes_filtrados_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )
    # Subconsulta para obtener la fecha del último movimiento
    ultimo_movimiento_fecha = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Agrega anotación y ordena por la fecha más reciente
    clientes_filtrados_qs = clientes_filtrados_qs.annotate(
        ultima_fecha=Subquery(ultimo_movimiento_fecha)
    ).order_by('-ultima_fecha')

    clientes_resultado = paginar_queryset(request, clientes_filtrados_qs, 'sin_actualizar')

    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()

    return render(request, 'clientes/clientes.html', {
        "clientes_completados": clientes_resultado,
        "view_type": "sin_actualizar",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": Cliente.objects.filter(estado_actual__nombre__iexact="actualizado", movimientos__estado__nombre__iexact="actualizado", movimientos__actualizado_por=usuario).distinct().count(),
        "count_sin_actualizar": clientes_filtrados_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_actualizados(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    movimiento_actualizado_por_mi = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk'),
        estado__nombre__iexact="actualizado",
        actualizado_por=usuario
    )
    fecha_movimiento_actualizado = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk'),
        estado__nombre__iexact="actualizado"
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    clientes_actualizados_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado"
    ).annotate(
        fue_actualizado_por_mi=Exists(movimiento_actualizado_por_mi),
        fecha_actualizado=Subquery(fecha_movimiento_actualizado)
    ).filter(fue_actualizado_por_mi=True).select_related('estado_actual')

    if search_query:
        clientes_actualizados_qs = clientes_actualizados_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )
    clientes_actualizados_qs = clientes_actualizados_qs.order_by('-fecha_actualizado')
    clientes_actualizados = paginar_queryset(request, clientes_actualizados_qs, 'actualizados')


    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()

    return render(request, 'clientes/clientes.html', {
        "clientes_actualizados": clientes_actualizados,
        "estado_reporte": EstadoReporte.objects.filter(estado="activo").exclude(nombre__in=["pendiente", "no localizado", "por localizar", "formulario sin respuesta"]),
        "view_type": "actualizados",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": clientes_actualizados_qs.count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@require_POST
def reportar_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    estado_id = request.POST.get("estado_reporte")
    nota_texto = request.POST.get("notas")
    estado_otro_nombre = request.POST.get("estado_otro")
    es_reenvio_directo = request.POST.get("accion_reenviar") == "1"

    if not cliente_id or not estado_id:
        messages.error(request, "Faltan datos del cliente o estado.")
        return redirect("clientes")

    cliente = get_object_or_404(Cliente, id=cliente_id)

    # Obtener o crear estado personalizado
    if estado_otro_nombre:
        estado = EstadoReporte.objects.create(
            nombre=estado_otro_nombre.strip(),
            creado_por=request.user,
            estado='activo',
            intentos=1,
            genera_movimiento=False
        )
    else:
        estado = get_object_or_404(EstadoReporte, id=estado_id)

    nombre_estado = estado.nombre.lower()

    # ------------------------ NO CONTESTÓ ------------------------
    if nombre_estado == "no contestó":
        cliente.sin_contestar += 1
        cliente.ultima_llamada_no_contesto = timezone.now()

        # Si aún no llega al límite, guardar historial sin movimiento
        if cliente.sin_contestar < estado.intentos:
            cliente.estado_actual = estado
            cliente.save()

            HistorialEstadoSinMovimiento.objects.create(
                cliente=cliente,
                estado=estado,
                actualizado_por=request.user,
                nota=nota_texto,
                genera_movimiento=False
            )

            messages.info(request, f"Intento de llamada {cliente.sin_contestar}/{estado.intentos}.")
            return redirect("clientes")

        # Si llega al límite, pasar a "por localizar"
        estado_por_localizar, _ = EstadoReporte.objects.get_or_create(
            nombre__iexact="por localizar",
            defaults={'nombre': 'Por localizar', 'creado_por': request.user, 'estado': 'activo'}
        )

        cliente.estado_actual = estado_por_localizar
        cliente.veces_contactado += 1
        cliente.sin_contestar = 0

        try:
            colector = User.objects.get(id=5)
            cliente.asignado_usuario = colector
        except User.DoesNotExist:
            pass

        cliente.save()

        # Guardar historial con genera_movimiento=True
        HistorialEstadoSinMovimiento.objects.create(
            cliente=cliente,
            estado=estado_por_localizar,
            actualizado_por=request.user,
            nota=nota_texto,
            genera_movimiento=True
        )

        movimiento = MovimientoEstado.objects.create(
            cliente=cliente,
            estado=estado_por_localizar,
            actualizado_por=request.user
        )

        if nota_texto:
            NotaMovimiento.objects.create(movimiento=movimiento, texto=nota_texto)

        messages.success(request, "Cliente enviado a colectores.")
        return redirect("clientes_colectores" if request.user.groups.filter(name="colector_group").exists() else "clientes")

    # ------------------------ FORMULARIO ENVIADO ------------------------
    if nombre_estado == "formulario enviado":
        cliente.formulario_sin_contestar += 1
        cliente.ultimo_envio_formulario = timezone.now()

        # Si aún no llega al límite, guardar historial sin movimiento
        if cliente.formulario_sin_contestar < estado.intentos:
            cliente.estado_actual = estado
            cliente.save()

            HistorialEstadoSinMovimiento.objects.create(
                cliente=cliente,
                estado=estado,
                actualizado_por=request.user,
                nota=nota_texto,
                genera_movimiento=False
            )

            messages.info(request, f"Formulario enviado. Envío {cliente.formulario_sin_contestar}/{estado.intentos}.")
            return redirect("clientes")

        # Si llega al límite, pasar a "formulario sin respuesta"
        estado_sin_respuesta = EstadoReporte.objects.filter(nombre__iexact="formulario sin respuesta").first()
        if estado_sin_respuesta:
            cliente.estado_actual = estado_sin_respuesta
            cliente.formulario_sin_contestar = 0
            cliente.save()

            HistorialEstadoSinMovimiento.objects.create(
                cliente=cliente,
                estado=estado_sin_respuesta,
                actualizado_por=request.user,
                nota=nota_texto,
                genera_movimiento=True
            )

            movimiento = MovimientoEstado.objects.create(
                cliente=cliente,
                estado=estado_sin_respuesta,
                actualizado_por=request.user
            )

            if nota_texto:
                NotaMovimiento.objects.create(movimiento=movimiento, texto=nota_texto)

            messages.success(request, "Cliente actualizado con estado Formulario sin respuesta.")
            return redirect("clientes")

        messages.info(request, f"Formulario enviado. Envío {cliente.formulario_sin_contestar}/{estado.intentos}.")
        return redirect("clientes")

    # ------------------------ OTROS ESTADOS ------------------------
    cliente.veces_contactado += 1
    cliente.sin_contestar = 0
    cliente.formulario_sin_contestar = 0
    cliente.estado_actual = estado
    cliente.save()

    if estado.genera_movimiento:
        movimiento = MovimientoEstado.objects.create(
            cliente=cliente,
            estado=estado,
            actualizado_por=request.user
        )
        if nota_texto:
            NotaMovimiento.objects.create(movimiento=movimiento, texto=nota_texto)
        messages.success(request, "Cliente actualizado exitosamente.")
    else:
        HistorialEstadoSinMovimiento.objects.create(
            cliente=cliente,
            estado=estado,
            actualizado_por=request.user,
            nota=nota_texto,
            genera_movimiento=False
        )
        messages.info(request, "Cliente registrado en seguimiento. Este cliente aún no se actualiza.")

    return redirect("clientes_colectores" if request.user.groups.filter(name="colector_group").exists() else "clientes")

@login_required
@require_POST
def actualizar_estado_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    cliente = get_object_or_404(Cliente, id=cliente_id)

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
    if not estado_actualizado:
        messages.error(request, "Estado 'Actualizado' no encontrado.")
        return redirect("clientes_colectores") if request.user.groups.filter(name="colector_group").exists() else redirect("clientes")

    cliente.veces_contactado += 1
    cliente.sin_contestar = 0
    cliente.estado_actual = estado_actualizado
    cliente.save()

    MovimientoEstado.objects.create(
        cliente=cliente,
        estado=estado_actualizado,
        actualizado_por=request.user,
        fecha_hora=timezone.now()
    )

    messages.success(request, "Cliente actualizado exitosamente.")
    return redirect("clientes_colectores") if request.user.groups.filter(name="colector_group").exists() else redirect("clientes")

@login_required
@require_POST
def crear_estado_reporte(request):
    nuevo_estado = request.POST.get("estado_otro", "").strip()

    if nuevo_estado:
        if EstadoReporte.objects.filter(nombre__iexact=nuevo_estado).exists():
            return JsonResponse({
                "success": False,
                "message": "Ya existe un estado con ese nombre.",
                "tags": "warning"
            })

        estado = EstadoReporte.objects.create(
            nombre=nuevo_estado,
            creado_por=request.user,
            estado="activo",
            intentos=1,
            genera_movimiento=False  # 👈🏽 Lo importante
        )

        return JsonResponse({
            "success": True,
            "id": estado.id,
            "nombre": estado.nombre,
            "message": "Estado creado exitosamente.",
            "tags": "success"
        })

    return JsonResponse({
        "success": False,
        "message": "Nombre de estado vacío.",
        "tags": "error"
    })

@login_required
def clientes_reportados(request):
    user = request.user
    grupos = user.groups.values_list('name', flat=True)
    search_query = request.GET.get('q', '').strip()

    # Filtrar clientes que tienen al menos un historial o un movimiento
    clientes_reportados_query = Cliente.objects.filter(
        Q(Exists(MovimientoEstado.objects.filter(cliente=OuterRef('pk')))) |
        Q(Exists(HistorialEstadoSinMovimiento.objects.filter(cliente=OuterRef('pk'))))
    ).prefetch_related(
        'movimientos__notas',
        'historial_sin_movimiento',
        'asignado_usuario',
        'asignado_inicial',
        'estado_actual'
    )

    # Filtro por grupo
    if "super_admin" in grupos or "admin_group" in grupos:
        clientes_filtrados = clientes_reportados_query

    elif "estandar_group" in grupos:
        clientes_filtrados = clientes_reportados_query.filter(
            asignado_inicial=user
        ).distinct()

    elif "colector_group" in grupos:
        clientes_filtrados = clientes_reportados_query.filter(
            Q(movimientos__actualizado_por=user) |
            Q(historial_sin_movimiento__actualizado_por=user)
        ).distinct()

    else:
        messages.error(request, "Acceso no permitido.")
        return redirect("inicio")

    # Búsqueda
    if search_query:
        clientes_filtrados = clientes_filtrados.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Paginación
    clientes_paginados = paginar_queryset(request, clientes_filtrados, 'reportados')

    # Procesar movimientos por cliente
    for cliente in clientes_paginados:
        movimientos_normales = [
            {"obj": m, "tipo": "con_movimiento"} for m in cliente.movimientos.all()
        ]
        movimientos_historial = [
            {"obj": h, "tipo": "sin_movimiento"}
            for h in cliente.historial_sin_movimiento.all()
            if not h.genera_movimiento
        ]

        cliente.todos_los_movimientos = sorted(
            movimientos_normales + movimientos_historial,
            key=lambda x: x["obj"].fecha_hora,
            reverse=True
        )

        cliente.ultimo_movimiento = cliente.movimientos.order_by('-fecha_hora').first()
        cliente.reportado_por = (
            cliente.todos_los_movimientos[0]["obj"].actualizado_por
            if cliente.todos_los_movimientos else None
        )

    return render(request, 'clientes/clientes_reportados.html', {
        "clientes": clientes_paginados,
        "view_type": "reportados",
        "search_query": search_query,
        "count_reportados": clientes_filtrados.count(),
    })

@login_required
def dashboard_reportes(request):
    if not request.user.groups.filter(name__in=['super_admin', 'admin_group']).exists():
        if request.user.groups.filter(name="colector_group").exists():
            messages.error(request, "Acceso no permitido.")
            return redirect("clientes_colectores")
        elif request.user.groups.filter(name="estandar_group").exists():
            messages.error(request, "Acceso no permitido.")
            return redirect("clientes")

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario_id = request.GET.get('usuario_id')

    grupo_estandar = Group.objects.get(name='estandar_group')
    grupo_colector = Group.objects.get(name='colector_group')

    movimientos = MovimientoEstado.objects.select_related('cliente', 'estado', 'actualizado_por')

    if fecha_inicio:
        fecha_inicio_dt = make_aware(datetime.combine(parse_date(fecha_inicio), datetime.min.time()))
        movimientos = movimientos.filter(fecha_hora__gte=fecha_inicio_dt)

    if fecha_fin:
        fecha_fin_dt = make_aware(datetime.combine(parse_date(fecha_fin), datetime.max.time()))
        movimientos = movimientos.filter(fecha_hora__lte=fecha_fin_dt)

    if usuario_id:
        movimientos = movimientos.filter(actualizado_por__id=usuario_id)

    movimientos_estandar = movimientos.filter(actualizado_por__groups=grupo_estandar)
    movimientos_colector = movimientos.filter(actualizado_por__groups=grupo_colector)

    clientes_estandar_qs = Cliente.objects.filter(movimientos__in=movimientos_estandar).distinct()
    clientes_colector_qs = Cliente.objects.filter(movimientos__in=movimientos_colector).distinct()

    total_reportes_estandar = movimientos_estandar.count()
    total_reportes_colector = movimientos_colector.count()

    clientes_estandar = clientes_estandar_qs.count()
    clientes_colector = clientes_colector_qs.count()

    clientes_totales = Cliente.objects.count()
    clientes_totales_colector = Cliente.objects.filter(movimientos__estado__nombre__iexact="por localizar").distinct().count()

    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
    clientes_sin_reportar_colector = Cliente.objects.filter(estado_actual=estado_por_localizar).count() if estado_por_localizar else 0
    clientes_no_asignados_colector = Cliente.objects.filter(
        estado_actual=estado_por_localizar, asignado_usuario__username="colector"
    ).count() if estado_por_localizar else 0

    clientes_asignados_estandar_qs = Cliente.objects.filter(asignado_inicial__groups=grupo_estandar)
    clientes_asignados_colector_qs = Cliente.objects.filter(asignado_usuario__groups=grupo_colector).exclude(asignado_usuario__username="colector")

    if usuario_id:
        clientes_asignados_estandar_qs = clientes_asignados_estandar_qs.filter(movimientos__actualizado_por__id=usuario_id)
        clientes_asignados_colector_qs = clientes_asignados_colector_qs.filter(movimientos__actualizado_por__id=usuario_id)

    if fecha_inicio:
        clientes_asignados_estandar_qs = clientes_asignados_estandar_qs.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)
        clientes_asignados_colector_qs = clientes_asignados_colector_qs.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)

    if fecha_fin:
        clientes_asignados_estandar_qs = clientes_asignados_estandar_qs.filter(movimientos__fecha_hora__lte=fecha_fin_dt)
        clientes_asignados_colector_qs = clientes_asignados_colector_qs.filter(movimientos__fecha_hora__lte=fecha_fin_dt)

    clientes_asignados_estandar = clientes_asignados_estandar_qs.distinct().count()
    clientes_asignados_colector = clientes_asignados_colector_qs.distinct().count()
    clientes_no_asignados_estandar = clientes_totales - clientes_asignados_estandar

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
    clientes_actualizados_estandar = clientes_actualizados_colector = 0

    if estado_actualizado:
        clientes_actualizados_estandar_qs = Cliente.objects.filter(
            movimientos__estado=estado_actualizado, movimientos__actualizado_por__groups=grupo_estandar
        )
        clientes_actualizados_colector_qs = Cliente.objects.filter(
            movimientos__estado=estado_actualizado, movimientos__actualizado_por__groups=grupo_colector
        ).exclude(asignado_usuario__username="colector")

        if usuario_id:
            clientes_actualizados_estandar_qs = clientes_actualizados_estandar_qs.filter(movimientos__actualizado_por__id=usuario_id)
            clientes_actualizados_colector_qs = clientes_actualizados_colector_qs.filter(movimientos__actualizado_por__id=usuario_id)

        if fecha_inicio:
            clientes_actualizados_estandar_qs = clientes_actualizados_estandar_qs.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)
            clientes_actualizados_colector_qs = clientes_actualizados_colector_qs.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)

        if fecha_fin:
            clientes_actualizados_estandar_qs = clientes_actualizados_estandar_qs.filter(movimientos__fecha_hora__lte=fecha_fin_dt)
            clientes_actualizados_colector_qs = clientes_actualizados_colector_qs.filter(movimientos__fecha_hora__lte=fecha_fin_dt)

        clientes_actualizados_estandar = clientes_actualizados_estandar_qs.distinct().count()
        clientes_actualizados_colector = clientes_actualizados_colector_qs.distinct().count()

    clientes_movidos_estandar = Cliente.objects.filter(movimientos__actualizado_por__groups=grupo_estandar).distinct()
    clientes_actualizados_por_estandar = Cliente.objects.filter(
        movimientos__estado=estado_actualizado, movimientos__actualizado_por__groups=grupo_estandar
    ).distinct()

    if usuario_id:
        clientes_movidos_estandar = clientes_movidos_estandar.filter(movimientos__actualizado_por__id=usuario_id)
        clientes_actualizados_por_estandar = clientes_actualizados_por_estandar.filter(movimientos__actualizado_por__id=usuario_id)

    if fecha_inicio:
        clientes_movidos_estandar = clientes_movidos_estandar.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)
        clientes_actualizados_por_estandar = clientes_actualizados_por_estandar.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)

    if fecha_fin:
        clientes_movidos_estandar = clientes_movidos_estandar.filter(movimientos__fecha_hora__lte=fecha_fin_dt)
        clientes_actualizados_por_estandar = clientes_actualizados_por_estandar.filter(movimientos__fecha_hora__lte=fecha_fin_dt)

    clientes_completados_estandar = clientes_movidos_estandar.exclude(
        id__in=clientes_actualizados_por_estandar.values_list('id', flat=True)
    ).count()

    clientes_completados_colector = Cliente.objects.filter(
        movimientos__estado__isnull=False, movimientos__actualizado_por__groups=grupo_colector
    ).exclude(
        movimientos__estado=estado_actualizado
    ).exclude(
        asignado_usuario__username="colector"
    )

    if usuario_id:
        clientes_completados_colector = clientes_completados_colector.filter(movimientos__actualizado_por__id=usuario_id)

    if fecha_inicio:
        clientes_completados_colector = clientes_completados_colector.filter(movimientos__fecha_hora__gte=fecha_inicio_dt)

    if fecha_fin:
        clientes_completados_colector = clientes_completados_colector.filter(movimientos__fecha_hora__lte=fecha_fin_dt)

    clientes_completados_colector = clientes_completados_colector.distinct().count()

    avance_total_estandar = clientes_actualizados_estandar + clientes_completados_estandar
    porcentaje_avance_estandar = round((avance_total_estandar / clientes_totales) * 100, 2) if clientes_totales > 0 else 0

    avance_total_colector = clientes_actualizados_colector + clientes_completados_colector
    porcentaje_avance_colector = round((avance_total_colector / clientes_totales_colector) * 100, 2) if clientes_totales_colector > 0 else 0

    reportes_estandar_por_estado = defaultdict(int)
    reportes_colector_por_estado = defaultdict(int)
    reportes_estandar_por_usuario = defaultdict(int)
    reportes_colector_por_usuario = defaultdict(int)
    actualizados_estandar_por_usuario = defaultdict(int)
    actualizados_colector_por_usuario = defaultdict(int)

    for m in movimientos_estandar:
        if m.estado:
            reportes_estandar_por_estado[m.estado.nombre] += 1
        if m.actualizado_por:
            nombre = m.actualizado_por.get_full_name()
            reportes_estandar_por_usuario[nombre] += 1
            if m.estado and m.estado.nombre.lower() == "actualizado":
                actualizados_estandar_por_usuario[nombre] += 1

    for m in movimientos_colector:
        if m.estado:
            reportes_colector_por_estado[m.estado.nombre] += 1
        if m.actualizado_por and m.actualizado_por.username != "colector":
            nombre = m.actualizado_por.get_full_name()
            reportes_colector_por_usuario[nombre] += 1
            if m.estado and m.estado.nombre.lower() == "actualizado":
                actualizados_colector_por_usuario[nombre] += 1

    usuarios = User.objects.filter(id__in=movimientos.values_list('actualizado_por__id', flat=True).distinct())

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "usuario_id": usuario_id,
        "usuarios": usuarios,
        "clientes_totales": clientes_totales,
        "clientes_totales_colector": clientes_totales_colector,
        "avance_total_estandar": avance_total_estandar,
        "porcentaje_avance_estandar": porcentaje_avance_estandar,
        "avance_total_colector": avance_total_colector,
        "porcentaje_avance_colector": porcentaje_avance_colector,
        "total_reportes_estandar": total_reportes_estandar,
        "clientes_estandar": clientes_estandar,
        "clientes_asignados_estandar": clientes_asignados_estandar,
        "clientes_no_asignados_estandar": clientes_no_asignados_estandar,
        "clientes_actualizados_estandar": clientes_actualizados_estandar,
        "clientes_completados_estandar": clientes_completados_estandar,
        "reportes_estandar_por_estado": dict(reportes_estandar_por_estado),
        "reportes_estandar_por_usuario": dict(reportes_estandar_por_usuario),
        "actualizados_estandar_por_usuario": dict(actualizados_estandar_por_usuario),
        "total_reportes_colector": total_reportes_colector,
        "clientes_colector": clientes_colector,
        "clientes_sin_reportar_colector": clientes_sin_reportar_colector,
        "clientes_asignados_colector": clientes_asignados_colector,
        "clientes_no_asignados_colector": clientes_no_asignados_colector,
        "clientes_actualizados_colector": clientes_actualizados_colector,
        "clientes_completados_colector": clientes_completados_colector,
        "reportes_colector_por_estado": dict(reportes_colector_por_estado),
        "reportes_colector_por_usuario": dict(reportes_colector_por_usuario),
        "actualizados_colector_por_usuario": dict(actualizados_colector_por_usuario),
    }

    return render(request, 'dashboard/dashboard.html', context)

@login_required
def clientes_sin_asignar_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()

    # Query base
    clientes_qs = Cliente.objects.filter(asignado_usuario__isnull=True)
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "sin_asignar")

    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs  = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.filter(
        movimientos__fecha_hora__date=hoy
    ).exclude(
        estado_actual__nombre__iexact="actualizado"
    ).distinct()

    usuarios_no_colectores = User.objects.exclude(groups__name="colector_group").filter(is_active=True)

    return render(request, "gestion/gestion.html", {
        "clientes_sin_asignar": clientes,
        "usuarios_no_colectores": usuarios_no_colectores,
        "view_type": "sin_asignar",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": Cliente.objects.exclude(
            estado_actual__nombre__iexact="actualizado"
        ).count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": clientes_qs.count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_actualizados_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Query base sin paginar (usada también para contadores)
    clientes_qs_base = Cliente.objects.filter(
        estado_actual=estado_actualizado,
        movimientos__estado=estado_actualizado
    ).distinct().prefetch_related('movimientos')

    # Filtro de búsqueda
    if search_query:
        clientes_qs_base = clientes_qs_base.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Aplicar paginación con prefijo 'actualizados'
    clientes = paginar_queryset(request, clientes_qs_base, "actualizados")

    # Contadores generales (no filtrados por usuario)
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, 'gestion/gestion.html', {
        "clientes_actualizados": clientes,
        "view_type": "actualizados",
        "search_query": search_query,
        "count_actualizados": clientes_qs_base.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_en_seguimiento_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")
    
    hoy = timezone.localdate()
    search_query = request.GET.get('q', '').strip()

    # Estados que definen a los clientes en seguimiento
    estados_seguimiento = EstadoReporte.objects.filter(genera_movimiento=False)\
        .exclude(nombre__iexact="pendiente")\
        .exclude(nombre__iexact="no contestó")

    # Subquery para obtener la última fecha en HistorialEstadoSinMovimiento
    ultimo_historial_qs = HistorialEstadoSinMovimiento.objects.filter(
        cliente=OuterRef("pk")
    ).order_by("-fecha_hora")

    # Query base para clientes en seguimiento con anotación de última fecha
    clientes_qs = Cliente.objects.filter(
        estado_actual__in=estados_seguimiento
    ).annotate(
        ultima_fecha_sin_movimiento=Subquery(
            ultimo_historial_qs.values("fecha_hora")[:1],
            output_field=DateTimeField()
        )
    )

    # Filtro por texto de búsqueda
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Paginación
    clientes = paginar_queryset(request, clientes_qs, 'seguimiento')

    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs = Cliente.objects.filter(
        estado_actual__in=estados_seguimiento
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, 'gestion/gestion.html', {
        "clientes_formulario_enviado": clientes,
        "view_type": "seguimiento",
        "search_query": search_query,
        "count_seguimiento": clientes_qs.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_pendientes_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()

    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    # Query principal: todos los clientes con estado pendiente (sin filtrar por usuario)
    clientes_qs = Cliente.objects.filter(estado_actual=estado_pendiente)

    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "pendientes")

    # Estados de seguimiento (sin generar movimiento, excepto 'pendiente' y 'no contestó')
    estados_seguimiento = EstadoReporte.objects.filter(genera_movimiento=False)\
        .exclude(nombre__iexact="pendiente")\
        .exclude(nombre__iexact="no contestó")

    clientes_seguimiento_qs = Cliente.objects.filter(estado_actual__in=estados_seguimiento)

    # Subqueries para detectar clientes actualizados pero cuyo primer estado no fue 'actualizado'
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef("pk")).order_by("fecha_hora")
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values("estado__nombre")[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, "gestion/gestion.html", {
        "clientes_pendientes": clientes,
        "view_type": "pendientes",
        "search_query": search_query,
        "count_pendientes": clientes_qs.count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_para_colectores_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()
    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()

    clientes_qs = Cliente.objects.filter(estado_actual=estado_por_localizar) if estado_por_localizar else Cliente.objects.none()

    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "colectores")


    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs  = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef("pk")).order_by("fecha_hora")
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values("estado__nombre")[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    usuarios_colectores = User.objects.filter(groups__name="colector_group", is_active=True)

    return render(request, "gestion/gestion.html", {
        "clientes_por_localizar": clientes,
        "usuarios_colectores": usuarios_colectores,
        "view_type": "colectores",
        "search_query": search_query,
        "count_colectores": clientes_qs.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
    })

login_required
def clientes_sin_actualizar_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    hoy = timezone.localdate()
    usuario = request.user

    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()

    # Estados definidos
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    # Obtener todos los estados que generan movimiento, excluyendo "actualizado"
    estados_con_movimiento = EstadoReporte.objects.filter(
        genera_movimiento=True
    ).exclude(nombre__iexact="actualizado")

    # Clientes cuyo estado actual es uno de los anteriores
    clientes_qs = Cliente.objects.filter(
        estado_actual__in=estados_con_movimiento
    ).select_related('estado_actual')

    # Subconsulta para obtener el primer estado del cliente
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_estado_nombre = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    # Subconsulta para obtener movimientos del cliente
    movimientos_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk'))

    # Anotaciones
    clientes_qs = clientes_qs.annotate(
        tiene_movimiento=Exists(movimientos_qs),
        primer_estado_nombre=primer_estado_nombre
    )

    # Filtro por búsqueda
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Anotaciones adicionales: fecha y usuario del último movimiento que coincide con el estado actual
    movimientos_compatibles = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk'),
        estado__nombre=OuterRef('estado_actual__nombre')
    ).order_by('-fecha_hora')

    clientes_qs = clientes_qs.annotate(
        ultima_fecha_movimiento=Subquery(
            movimientos_compatibles.values('fecha_hora')[:1],
            output_field=DateTimeField()
        ),
        usuario_movimiento=Subquery(
            movimientos_compatibles.values('actualizado_por__username')[:1]
        )
    ).order_by('-ultima_fecha_movimiento')

    clientes_resultado = paginar_queryset(request, clientes_qs, 'sin_actualizar')

    # Contadores
    clientes_seguimiento_qs = Cliente.objects.filter(estado_actual__in=estados_con_movimiento)
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = clientes_qs.filter(
        ultima_fecha_movimiento__date=hoy
    )

    return render(request, 'clientes/clientes.html', {
        "clientes_completados": clientes_resultado,
        "view_type": "sin_actualizar_global",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado",
            movimientos__actualizado_por=usuario
        ).distinct().count(),
        "count_sin_actualizar": clientes_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_todos_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()

    clientes_qs = Cliente.objects.select_related("estado_actual", "asignado_usuario")
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "todos")

    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs  = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef("pk")).order_by("fecha_hora")
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values("estado__nombre")[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, "gestion/gestion.html", {
        "clientes_todos": clientes,
        "view_type": "todos",
        "search_query": search_query,
        "count_todos": clientes_qs.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
@require_POST
def asignar_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    usuario_id = request.POST.get("usuario_id")

    cliente = get_object_or_404(Cliente, id=cliente_id)
    usuario = get_object_or_404(User, id=usuario_id)

    # 👇 SOLO registrar la primera vez
    if cliente.asignado_inicial is None:
        cliente.asignado_inicial = usuario

    cliente.asignado_usuario = usuario
    cliente.save()

    messages.success(request, f"Cliente asignado exitosamente a {usuario.get_full_name()}.")
    return redirect("gestion")

@login_required
@require_POST
def asignacion_por_cantidad(request):
    cantidad = int(request.POST.get('cantidad', 0))
    usuario_id = request.POST.get('usuario_id')

    if cantidad <= 0 or not usuario_id:
        messages.error(request, "Debe ingresar una cantidad válida y seleccionar un usuario.")
        return redirect('gestion')

    usuario = get_object_or_404(User, id=usuario_id)

    # No se permite asignar a usuarios del grupo colector
    if usuario.groups.filter(name="colector_group").exists():
        messages.error(request, "No se puede asignar clientes a usuarios del grupo 'colector'.")
        return redirect('gestion')

    # Obtener todos los clientes no asignados
    todos_no_asignados = Cliente.objects.filter(asignado_usuario__isnull=True).order_by('id')
    total_disponibles = todos_no_asignados.count()

    if total_disponibles == 0:
        messages.warning(request, "No hay clientes disponibles para asignar.")
        return redirect('gestion')

    if total_disponibles < cantidad:
        messages.warning(request, f"Solo hay {total_disponibles} cliente(s) sin asignar. Ajuste la cantidad.")
        return redirect('gestion')

    # Obtener solo los primeros `cantidad`
    clientes_a_asignar = todos_no_asignados[:cantidad]

    for cliente in clientes_a_asignar:
        # ✅ Solo asignar el inicial si aún no ha sido asignado antes
        if cliente.asignado_inicial is None:
            cliente.asignado_inicial = usuario

        cliente.asignado_usuario = usuario
        cliente.save()

    messages.success(request, f"{clientes_a_asignar.count()} clientes asignados a {usuario.get_full_name()}.")
    return redirect('gestion')

@login_required
@require_POST
def reasignar_cliente_colector(request):
    cliente_id = request.POST.get("cliente_id")
    usuario_id = request.POST.get("usuario_id")

    cliente = get_object_or_404(Cliente, id=cliente_id)
    nuevo_usuario = get_object_or_404(User, id=usuario_id)

    # ✅ Solo se permite asignar a usuarios del grupo colector_group
    if not nuevo_usuario.groups.filter(name="colector_group").exists():
        messages.error(request, "Solo se puede asignar a usuarios del grupo 'colector_group'.")
        return redirect("gestion")

    cliente.asignado_usuario = nuevo_usuario
    cliente.save()

    messages.success(request, f"Cliente asignado exitosamente al colector {nuevo_usuario.get_full_name()}.")
    return redirect("gestion")

@login_required
def clientes_colectores(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        if request.user.groups.filter(name="estandar_group").exists():
            return redirect("clientes")
        elif request.user.groups.filter(name__in=["super_admin", "admin_group"]).exists():
            return redirect("gestion")
        else:
            return redirect("login")

    # Estados necesarios
    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Clientes con estado "por localizar" asignados al usuario
    clientes_por_localizar_asignados = Cliente.objects.none()
    clientes_por_localizar_filtrados = []

    if estado_por_localizar:
        clientes_por_localizar_asignados = Cliente.objects.filter(
            movimientos__estado=estado_por_localizar,
            asignado_usuario=request.user
        ).distinct().prefetch_related('movimientos')

        for cliente in clientes_por_localizar_asignados:
            movimientos_ordenados = sorted(cliente.movimientos.all(), key=lambda m: m.fecha_hora, reverse=True)
            if movimientos_ordenados and movimientos_ordenados[0].estado == estado_por_localizar:
                cliente.ultimo_movimiento = movimientos_ordenados[0]
                clientes_por_localizar_filtrados.append(cliente)

    # Movimientos con estado actualizado por el usuario
    movimientos_actualizados_por_usuario = MovimientoEstado.objects.none()
    if estado_actualizado:
        movimientos_actualizados_por_usuario = MovimientoEstado.objects.filter(
            estado=estado_actualizado,
            actualizado_por=request.user
        ).select_related('cliente', 'estado')

    # Estados visibles en el modal
    estados_incluir_modal = ["Actualizado", "Se negó", "No localizado", "Liquidada"]
    estados_visibles = EstadoReporte.objects.filter(nombre__in=estados_incluir_modal)

    # Clientes completados (reportados pero no actualizados)
    clientes_completados = Cliente.objects.filter(
        movimientos__actualizado_por=request.user
    ).distinct()

    if estado_actualizado:
        clientes_completados = clientes_completados.exclude(estado_actual=estado_actualizado)

    movimientos_prefetch = Prefetch(
        'movimientos',
        queryset=MovimientoEstado.objects.select_related('actualizado_por').order_by('-fecha_hora'),
        to_attr='movimientos_ordenados'
    )
    clientes_completados = clientes_completados.prefetch_related(movimientos_prefetch)

    for cliente in clientes_completados:
        cliente.ultimo_movimiento = cliente.movimientos_ordenados[0] if cliente.movimientos_ordenados else None

    return render(request, "clientes/clientes_colectores.html", {
        "clientes_por_localizar_filtrados": clientes_por_localizar_filtrados,
        "movimientos_actualizados_por_usuario": movimientos_actualizados_por_usuario,
        "clientes_completados": clientes_completados,
        "estado_reporte": estados_visibles
    })


@login_required
def clientes_colectores_pendientes(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    usuario = request.user
    search_query = request.GET.get("q", "").strip()
    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Subquery para obtener el último estado (por fecha) de cada cliente
    ult_estado_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Clientes asignados al usuario y con último estado 'por localizar'
    clientes_por_localizar_qs = Cliente.objects.filter(
        asignado_usuario=usuario
    ).annotate(
        ultimo_estado=Subquery(ult_estado_subquery),
        ultima_fecha=Subquery(ult_fecha_subquery)
    ).filter(
        ultimo_estado__iexact="por localizar"
    ).order_by('-ultima_fecha')

    if search_query:
        clientes_por_localizar_qs = clientes_por_localizar_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes_por_localizar = paginar_queryset(request, clientes_por_localizar_qs, 'pendientes')

    return render(request, "clientes/clientes_colectores.html", {
        "clientes": clientes_por_localizar,
        "view_type": "pendientes",
        "search_query": search_query,
        "count_pendientes": clientes_por_localizar_qs.count(),
        "count_completados": Cliente.objects.filter(movimientos__actualizado_por=usuario).exclude(estado_actual=estado_actualizado).distinct().count(),
        "count_actualizados": MovimientoEstado.objects.filter(estado=estado_actualizado, actualizado_por=usuario).count(),
        "estado_reporte": EstadoReporte.objects.filter(nombre__in=["Actualizado", "Se negó", "No localizado", "Liquidada"])
    })

@login_required
def clientes_colectores_completados(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    usuario = request.user
    search_query = request.GET.get("q", "").strip()

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Subqueries para última fecha y último estado
    ult_estado_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Clientes completados (reportados pero no actualizados)
    clientes_completados_qs = Cliente.objects.filter(
        movimientos__actualizado_por=usuario
    ).exclude(
        estado_actual=estado_actualizado
    ).annotate(
        ultima_fecha=Subquery(ult_fecha_subquery),
        ultimo_estado=Subquery(ult_estado_subquery)
    ).order_by('-ultima_fecha').distinct()

    # Prefetch de movimientos ordenados
    movimientos_prefetch = Prefetch(
        'movimientos',
        queryset=MovimientoEstado.objects.select_related('estado').order_by('-fecha_hora'),
        to_attr='movimientos_ordenados'
    )
    clientes_completados_qs = clientes_completados_qs.prefetch_related(movimientos_prefetch)

    if search_query:
        clientes_completados_qs = clientes_completados_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes_completados = paginar_queryset(request, clientes_completados_qs, 'completados')

    # Asignar último movimiento para visualización
    for cliente in clientes_completados:
        cliente.ultimo_movimiento = cliente.movimientos_ordenados[0] if cliente.movimientos_ordenados else None

    # Contadores coherentes
    clientes_pendientes_qs = Cliente.objects.filter(
        asignado_usuario=usuario
    ).annotate(
        ultimo_estado=Subquery(ult_estado_subquery)
    ).filter(
        ultimo_estado__iexact="por localizar"
    )

    return render(request, "clientes/clientes_colectores.html", {
        "clientes": clientes_completados,
        "view_type": "completados",
        "search_query": search_query,
        "count_pendientes": clientes_pendientes_qs.count(),
        "count_completados": clientes_completados_qs.count(),
        "count_actualizados": MovimientoEstado.objects.filter(estado=estado_actualizado, actualizado_por=usuario).count(),
        "estado_reporte": EstadoReporte.objects.filter(nombre__in=["Actualizado", "Se negó", "No localizado", "Liquidada"])
    })

@login_required
def clientes_colectores_actualizados(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    usuario = request.user
    search_query = request.GET.get("q", "").strip()

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    movimientos_actualizados_qs = MovimientoEstado.objects.filter(
        estado=estado_actualizado,
        actualizado_por=usuario
    ).select_related("cliente")

    if search_query:
        movimientos_actualizados_qs = movimientos_actualizados_qs.filter(
            Q(cliente__nombre_cliente__icontains=search_query) |
            Q(cliente__numero_cliente__icontains=search_query) |
            Q(cliente__contacto_cliente__icontains=search_query)
        )

    movimientos_actualizados = paginar_queryset(request, movimientos_actualizados_qs.order_by('-fecha_hora'), 'actualizados')

    # Subqueries para contar pendientes
    ult_estado_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    clientes_pendientes_qs = Cliente.objects.filter(
        asignado_usuario=usuario
    ).annotate(
        ultimo_estado=Subquery(ult_estado_subquery),
        ultima_fecha=Subquery(ult_fecha_subquery)
    ).filter(
        ultimo_estado__iexact="por localizar"
    )

    clientes_completados_qs = Cliente.objects.filter(
        movimientos__actualizado_por=usuario
    ).exclude(estado_actual=estado_actualizado).distinct()

    return render(request, "clientes/clientes_colectores.html", {
        "movimientos_actualizados": movimientos_actualizados,
        "view_type": "actualizados",
        "search_query": search_query,
        "count_pendientes": clientes_pendientes_qs.count(),
        "count_completados": clientes_completados_qs.count(),
        "count_actualizados": movimientos_actualizados_qs.count(),
        "estado_reporte": EstadoReporte.objects.filter(nombre__in=["Actualizado", "Se negó", "No localizado", "Liquidada"])
    })

@login_required
@require_POST
def importar_clientes(request):
    archivo = request.FILES.get("archivo_excel")
    if not archivo:
        messages.error(request, "Debe subir un archivo Excel.")
        return redirect("gestion")

    try:
        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active

        filas_insertadas = 0
        filas_actualizadas = 0
        errores = []

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            try:
                numero, nombre, direccion, contacto, telefono, telefono_dos, correo = fila[:7]
            except ValueError:
                errores.append(f"Fila {i}: Número incorrecto de columnas.")
                continue

            if not (numero and nombre):
                errores.append(f"Fila {i}: número o nombre faltante.")
                continue

            cliente = Cliente.objects.filter(numero_cliente=numero).first()
            estado = EstadoReporte.objects.filter(id=1).first()  # Estado "Pendiente"

            if not estado:
                errores.append(f"Fila {i}: No se encontró el estado con ID 1 (Pendiente).")
                continue

            if cliente:
                cliente.nombre_cliente = nombre
                cliente.contacto_cliente = contacto
                cliente.telefono_cliente = str(telefono)
                cliente.telefono_dos = str(telefono_dos) if telefono_dos else None
                cliente.correo = correo or None
                cliente.direccion = direccion or None
                cliente.estado_actual = estado
                cliente.save()
                filas_actualizadas += 1
            else:
                Cliente.objects.create(
                    numero_cliente=numero,
                    nombre_cliente=nombre,
                    contacto_cliente=contacto,
                    telefono_cliente=str(telefono),
                    telefono_dos=str(telefono_dos) if telefono_dos else None,
                    correo=correo or None,
                    direccion=direccion or None,
                    estado_actual=estado
                )
                filas_insertadas += 1

        msg = f"Importación completada: {filas_insertadas} insertado(s), {filas_actualizadas} actualizado(s)."
        if errores:
            msg += f" {len(errores)} error(es) encontrados. Revisa el archivo."

        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect("gestion")

@login_required
@require_POST
def exportar_clientes(request):
    filtro = request.POST.get("filtro_exportacion")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_fin")

    queryset = Cliente.objects.all().select_related("estado_actual", "asignado_usuario")
    zona_honduras = pytz.timezone("America/Tegucigalpa")

    if filtro == "actualizados":
        estado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
        if estado:
            queryset = queryset.filter(movimientos__estado=estado).distinct()

        if fecha_inicio:
            queryset = queryset.filter(movimientos__fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(movimientos__fecha_hora__date__lte=fecha_fin)

    elif filtro == "seguimiento":
        estados = EstadoReporte.objects.filter(genera_movimiento=False).exclude(nombre__iexact="pendiente")
        queryset = queryset.filter(estado_actual__in=estados)

        ultimo_historial_qs = HistorialEstadoSinMovimiento.objects.filter(
            cliente=OuterRef("pk")
        ).order_by("-fecha_hora")

        queryset = queryset.annotate(
            ultima_fecha_sin_movimiento=Subquery(
                ultimo_historial_qs.values("fecha_hora")[:1],
                output_field=DateTimeField()
            )
        )

        if fecha_inicio:
            queryset = queryset.filter(
                Q(ultimo_envio_formulario__date__gte=fecha_inicio) |
                Q(ultima_fecha_sin_movimiento__date__gte=fecha_inicio)
            )
        if fecha_fin:
            queryset = queryset.filter(
                Q(ultimo_envio_formulario__date__lte=fecha_fin) |
                Q(ultima_fecha_sin_movimiento__date__lte=fecha_fin)
            )

    elif filtro == "colectores":
        estado = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
        if estado:
            queryset = queryset.filter(estado_actual=estado)

        if fecha_inicio:
            queryset = queryset.filter(movimientos__fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(movimientos__fecha_hora__date__lte=fecha_fin)


    elif filtro == "pendientes":
        estado = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
        if estado:
            queryset = queryset.filter(estado_actual=estado)

        if fecha_inicio:
            queryset = queryset.filter(movimientos__fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(movimientos__fecha_hora__date__lte=fecha_fin)

    elif filtro == "completados":
        estados_con_movimiento = EstadoReporte.objects.filter(
            genera_movimiento=True
        ).exclude(nombre__iexact="actualizado")

        movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'))
        primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
        primer_estado_nombre = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

        movimientos_compatibles = MovimientoEstado.objects.filter(
            cliente=OuterRef('pk'),
            estado__nombre=OuterRef('estado_actual__nombre')
        ).order_by('-fecha_hora')

        queryset = Cliente.objects.filter(
            estado_actual__in=estados_con_movimiento
        ).annotate(
            tiene_movimiento=Exists(movimientos_usuario),
            primer_estado_nombre=primer_estado_nombre,
            ultima_fecha_movimiento=Subquery(
                movimientos_compatibles.values('fecha_hora')[:1],
                output_field=DateTimeField()
            ),
            usuario_movimiento=Subquery(
                movimientos_compatibles.values('actualizado_por__username')[:1]
            )
        ).order_by('-ultima_fecha_movimiento')

        if fecha_inicio:
            queryset = queryset.filter(ultima_fecha_movimiento__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(ultima_fecha_movimiento__date__lte=fecha_fin)

    elif filtro == "todos":
        queryset = Cliente.objects.select_related("estado_actual", "asignado_usuario").all()
        # Puedes filtrar por fechas si tienes un campo de fecha en Cliente
        # if fecha_inicio:
        #     queryset = queryset.filter(fecha_creacion__date__gte=fecha_inicio)
        # if fecha_fin:
        #     queryset = queryset.filter(fecha_creacion__date__lte=fecha_fin)

    # Crear el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Exportados"
    ws.append(["#", "Cliente", "Nombre", "Contacto", "Estado", "Usuario", "Fecha"])

    for i, cliente in enumerate(queryset, start=1):
        fecha = ""
        usuario = cliente.asignado_usuario.get_full_name() if cliente.asignado_usuario else "Sin asignar"

        if filtro == "seguimiento":
            if hasattr(cliente, "ultimo_envio_formulario") and cliente.ultimo_envio_formulario:
                fecha = timezone.localtime(cliente.ultimo_envio_formulario, zona_honduras).strftime("%d/%m/%Y %H:%M")
            elif hasattr(cliente, "ultima_fecha_sin_movimiento") and cliente.ultima_fecha_sin_movimiento:
                fecha = timezone.localtime(cliente.ultima_fecha_sin_movimiento, zona_honduras).strftime("%d/%m/%Y %H:%M")

        elif filtro == "completados":
            fecha = timezone.localtime(cliente.ultima_fecha_movimiento, zona_honduras).strftime("%d/%m/%Y %H:%M") if cliente.ultima_fecha_movimiento else ""
            usuario = cliente.usuario_movimiento if cliente.usuario_movimiento else "Sin asignar"

        elif filtro != "todos":
            ultimo_mov = cliente.movimientos.order_by("-fecha_hora").first()
            if ultimo_mov:
                fecha = timezone.localtime(ultimo_mov.fecha_hora, zona_honduras).strftime("%d/%m/%Y %H:%M")

        ws.append([
            i,
            cliente.numero_cliente,
            cliente.nombre_cliente,
            cliente.contacto_cliente,
            cliente.estado_actual.nombre if cliente.estado_actual else "",
            usuario,
            fecha,
        ])

    # Generar nombre dinámico del archivo
    nombre_base = f"Clientes_{filtro}"
    if fecha_inicio and fecha_fin:
        nombre_archivo = f"{nombre_base}_{fecha_inicio}_a_{fecha_fin}.xlsx"
    elif fecha_inicio:
        nombre_archivo = f"{nombre_base}_desde_{fecha_inicio}.xlsx"
    elif fecha_fin:
        nombre_archivo = f"{nombre_base}_hasta_{fecha_fin}.xlsx"
    else:
        nombre_archivo = f"{nombre_base}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    wb.save(response)
    return response